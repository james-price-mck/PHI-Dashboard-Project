"""
Build the static dashboard JSON from several sources:

1. APRA "Quarterly Private Health Insurance Membership Trends" xlsx  -> national & jurisdictional
   hospital / general coverage and insured persons (https://www.apra.gov.au/quarterly-private-health-insurance-statistics).
2. APRA "Quarterly Private Health Insurance Membership and Benefits" xlsx  -> AgeCohort_HT
   (hospital insured persons by 5-year age band, quarterly, since 2007).
3. ABS ERP_Q via SDMX (see etl/fetch_abs_erp.py)  -> national 5-year age band population
   denominators, quarterly.
4. Department of Health, Disability and Ageing "Private Health Insurance Reform Data
   Quarterly Trends Report" xlsx  -> sheet "4 HT by Product Tier", second block
   (hospital-treatment insured persons by Gold/Silver/Bronze/Basic/Other, quarterly,
   from Jun-2019).
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import asdict, dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

# HT % coverage: header row 22 (0-based index 22) = row 23 in Excel — verified Dec 2025 file.
HT_COVERAGE_HEADER = 22
# GT % coverage: table starts later in sheet.
GT_COVERAGE_HEADER = 30

HT_SHEET = "HT % coverage"
GT_SHEET = "GT % coverage"
AGE_COHORT_SHEET = "AgeCohort_HT"
TIER_SHEET = "4 HT by Product Tier"

# APRA AgeCohort_HT uses numeric "Age" values that are the lower bound of a 5-year band
# (0 -> 0-4, 5 -> 5-9, ...). ABS ERP_Q only publishes bands up to A80 ("80 and over"),
# so we collapse APRA's 80/85/90/95 bands into a single "80+" band to match.
APRA_AGE_TO_BAND: dict[int, str] = {
    0: "0-4",
    5: "5-9",
    10: "10-14",
    15: "15-19",
    20: "20-24",
    25: "25-29",
    30: "30-34",
    35: "35-39",
    40: "40-44",
    45: "45-49",
    50: "50-54",
    55: "55-59",
    60: "60-64",
    65: "65-69",
    70: "70-74",
    75: "75-79",
    80: "80+",
    85: "80+",
    90: "80+",
    95: "80+",
}
AGE_BAND_ORDER = [
    "0-4",
    "5-9",
    "10-14",
    "15-19",
    "20-24",
    "25-29",
    "30-34",
    "35-39",
    "40-44",
    "45-49",
    "50-54",
    "55-59",
    "60-64",
    "65-69",
    "70-74",
    "75-79",
    "80+",
]
MONTH_END_TO_QUARTER: dict[str, str] = {
    "Mar": "03-31",
    "Jun": "06-30",
    "Sep": "09-30",
    "Dec": "12-31",
}

# Tier reforms effective; Gold/Silver/Bronze/Basic product tiers rolled out from 1 April 2019
# and became mandatory on 1 April 2020. All dashboard series are cut to this floor so the
# time window is internally consistent with the tier-mix view.
SERIES_START_ISO = "2019-04-01"

# APRA: reporting framework / collection change — annotate charts (see APRA private health insurance statistics)
APRA_METHODOLOGY_CHANGE = "2023-07-01"
TIER_REFORM_ISO = "2019-04-01"

# Tier taxonomy — order matters for the stacked chart (bottom to top).
TIER_ORDER = ["gold", "silver", "bronze", "basic", "other"]
TIER_LABEL = {
    "gold": "Gold",
    "silver": "Silver",
    "bronze": "Bronze",
    "basic": "Basic",
    "other": "Legacy (pre-reform)",
}

# Column index of each tier's "Total Policies" / "Total Persons" summary column in
# sheet "4 HT by Product Tier" of the DoH workbook (verified Dec 2025 edition).
TIER_TOTAL_COLS: dict[str, int] = {
    "gold": 7,
    "silver": 14,
    "bronze": 21,
    "basic": 28,
    "other": 35,
    "grand_total": 42,
}

STATE_KEYS_HT = [
    "NSW & ACT",
    "VIC",
    "QLD",
    "SA & NT*",
    "WA",
    "TAS",
    "ACT",
    "NT**",
    "AUST",
]
@dataclass
class ReconciliationResult:
    passed: bool
    metric: str
    expected: float | None
    actual: float | None
    abs_diff: float | None
    tolerance: float
    message: str


def _q_iso(ts: Any) -> str:
    if isinstance(ts, datetime):
        return ts.date().isoformat()
    if isinstance(ts, date):
        return ts.isoformat()
    return str(ts)[:10]


def load_ht_gt(path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    ht = pd.read_excel(path, sheet_name=HT_SHEET, header=HT_COVERAGE_HEADER)
    gt = pd.read_excel(path, sheet_name=GT_SHEET, header=GT_COVERAGE_HEADER)
    return ht, gt


def build_national_quarterly_simple(ht: pd.DataFrame, gt: pd.DataFrame) -> list[dict[str, Any]]:
    h = ht.dropna(subset=["Quarter"]).copy()
    g = gt.dropna(subset=["Quarter"]).copy()
    h["Quarter"] = pd.to_datetime(h["Quarter"])
    g["Quarter"] = pd.to_datetime(g["Quarter"])
    h = h[h["Quarter"] >= SERIES_START_ISO]
    g = g[g["Quarter"] >= SERIES_START_ISO]
    hq = h.set_index("Quarter").sort_index()
    gq = g.set_index("Quarter").sort_index()
    common = sorted(hq.index.intersection(gq.index))
    rows: list[dict[str, Any]] = []
    for q in common:
        rh = hq.loc[q]
        rg = gq.loc[q]
        if isinstance(rh, pd.DataFrame):
            rh = rh.iloc[0]
        if isinstance(rg, pd.DataFrame):
            rg = rg.iloc[0]
        h_rate = float(rh["AUST"]) if pd.notna(rh["AUST"]) else None
        h_pers = float(rh["AUST.1"]) if pd.notna(rh.get("AUST.1")) else None
        h_erp = float(rh["AUST.2"]) if pd.notna(rh.get("AUST.2")) else None
        g_rate = float(rg["AUST"]) if pd.notna(rg["AUST"]) else None
        g_pers = float(rg["AUST.1"]) if pd.notna(rg.get("AUST.1")) else None
        g_erp = float(rg["AUST.2"]) if pd.notna(rg.get("AUST.2")) else None
        rows.append(
            {
                "quarter": _q_iso(q),
                "hospital_treatment": {
                    "insured_persons": h_pers,
                    "share_of_population": h_rate,
                    "population_denominator": h_erp,
                },
                "general_treatment": {
                    "insured_persons": g_pers,
                    "share_of_population": g_rate,
                    "population_denominator": g_erp,
                },
            }
        )
    return rows


def build_state_quarterly_clean(ht: pd.DataFrame) -> list[dict[str, Any]]:
    h = ht.dropna(subset=["Quarter"]).copy()
    h["Quarter"] = pd.to_datetime(h["Quarter"])
    h = h[h["Quarter"] >= SERIES_START_ISO]
    labels = [x for x in STATE_KEYS_HT if x != "AUST"]
    out: list[dict[str, Any]] = []
    for _, r in h.iterrows():
        q = r["Quarter"]
        jur: dict[str, Any] = {}
        for label in labels:
            jcode = label.replace(" & ", "_").replace("*", "").replace("**", "")
            if label == "SA & NT*":
                sh = float(r[label]) if pd.notna(r.get(label)) else None
                # APRA: combined % in first block; persons split under SA / NT in the second
                # block, populations under SA.1 / NT.1 (third block). No single combined
                # "insured persons" column — derive a consistent count from % × combined pop.
                p_sa = r.get("SA")
                p_nt = r.get("NT")
                pop_sa = r.get("SA.1")
                pop_nt = r.get("NT.1")
                pop = None
                if pd.notna(pop_sa) and pd.notna(pop_nt):
                    pop = float(pop_sa) + float(pop_nt)
                p = None
                if sh is not None and pop is not None:
                    p = sh * pop
                elif pd.notna(p_sa) and pd.notna(p_nt):
                    p = float(p_sa) + float(p_nt)
                if sh is not None or p is not None or pop is not None:
                    jur["SA_NT"] = {
                        "share_of_population": sh,
                        "insured_persons": p,
                        "population_denominator": pop,
                    }
                continue
            k1 = f"{label}.1"
            k2 = f"{label}.2"
            sh = float(r[label]) if label in h.columns and pd.notna(r.get(label)) else None
            p = float(r[k1]) if k1 in h.columns and pd.notna(r.get(k1)) else None
            pop = float(r[k2]) if k2 in h.columns and pd.notna(r.get(k2)) else None
            if sh is not None or p is not None or pop is not None:
                jur[jcode] = {
                    "share_of_population": sh,
                    "insured_persons": p,
                    "population_denominator": pop,
                }
        out.append({"quarter": _q_iso(q), "jurisdictions": jur})
    return out


def load_age_cohort_ht(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=AGE_COHORT_SHEET, header=0)
    # Two header rows in the source; drop the repeated one.
    df = df[df["Year"] != "Year"].copy()
    df["Year"] = pd.to_numeric(df["Year"], errors="coerce").astype("Int64")
    df["Age"] = pd.to_numeric(df["Age"], errors="coerce").astype("Int64")
    df["InsuredPersons"] = pd.to_numeric(df["InsuredPersons"], errors="coerce")
    df = df.dropna(subset=["Year", "Age", "MonthEnd", "InsuredPersons"])
    return df


def load_abs_erp(path: Path) -> pd.DataFrame:
    """Tidy CSV written by etl/fetch_abs_erp.py: quarter (YYYY-Qn), age_band, population."""
    df = pd.read_csv(path)
    df["population"] = pd.to_numeric(df["population"], errors="coerce")
    return df.dropna(subset=["quarter", "age_band", "population"])


def _quarter_from_year_monthend(year: int, month_end: str) -> str | None:
    md = MONTH_END_TO_QUARTER.get(month_end)
    if not md:
        return None
    return f"{int(year)}-{md}"


def _abs_quarter_to_iso(q: str) -> str:
    y, qn = q.split("-Q")
    end = {"1": "03-31", "2": "06-30", "3": "09-30", "4": "12-31"}[qn]
    return f"{y}-{end}"


def build_age_coverage_quarterly(
    age_df: pd.DataFrame,
    erp_df: pd.DataFrame,
) -> tuple[list[dict[str, Any]], str | None]:
    """National hospital-cover rate by 5-year age band, quarterly, joined to ABS ERP.

    Returns (records, stale_denominator_notice). If APRA data extends beyond the most
    recent ABS ERP quarter, the last available ABS ERP quarter is reused and a notice
    describing that is returned (used for the About panel).
    """
    # National insured persons by APRA band per quarter (sum of state & gender).
    a = age_df.copy()
    a["band"] = a["Age"].map(APRA_AGE_TO_BAND)
    a = a.dropna(subset=["band"])
    a["quarter"] = [
        _quarter_from_year_monthend(int(y), str(m))
        for y, m in zip(a["Year"], a["MonthEnd"])
    ]
    a = a.dropna(subset=["quarter"])
    a = a[a["quarter"] >= SERIES_START_ISO]
    insured = (
        a.groupby(["quarter", "band"], as_index=False)["InsuredPersons"]
        .sum()
        .rename(columns={"InsuredPersons": "insured_persons"})
    )

    # ABS ERP — quarter labels are 'YYYY-Qn'; translate to ISO quarter-end for join.
    e = erp_df.copy()
    e["quarter_iso"] = e["quarter"].map(_abs_quarter_to_iso)
    e = e.rename(columns={"age_band": "band", "population": "erp"})
    e = e[["quarter_iso", "band", "erp"]].rename(columns={"quarter_iso": "quarter"})

    latest_abs = e["quarter"].max() if not e.empty else None

    # Left-join; for APRA quarters beyond the latest ABS quarter, fall back to the
    # latest ABS age-band distribution (a conservative hold-forward).
    merged = insured.merge(e, on=["quarter", "band"], how="left")
    missing_mask = merged["erp"].isna()
    if missing_mask.any() and latest_abs is not None:
        fallback = e[e["quarter"] == latest_abs].set_index("band")["erp"]
        merged.loc[missing_mask, "erp"] = merged.loc[missing_mask, "band"].map(fallback)

    merged["coverage_rate"] = merged["insured_persons"] / merged["erp"]

    out: list[dict[str, Any]] = []
    for q, grp in merged.groupby("quarter"):
        bands: dict[str, dict[str, float]] = {}
        for _, r in grp.iterrows():
            if pd.isna(r["erp"]) or r["erp"] <= 0:
                continue
            bands[str(r["band"])] = {
                "insured_persons": float(r["insured_persons"]),
                "population": float(r["erp"]),
                "coverage_rate": float(r["coverage_rate"]),
            }
        if bands:
            out.append({"quarter": str(q), "bands": bands})

    out.sort(key=lambda x: x["quarter"])

    notice = None
    if latest_abs is not None and out:
        last_apra = out[-1]["quarter"]
        if last_apra > latest_abs:
            notice = (
                f"ABS ERP by 5-year age band is published up to {latest_abs}; "
                f"for {last_apra} the most recent ABS age-band populations are held forward. "
                "Revise on each ABS release."
            )
    return out, notice


def _num(v: Any) -> float | None:
    """Parse a DoH tier-sheet cell into a float, or None for blanks / suppressed values.

    The workbook uses:
      - '' / None for genuinely empty cells (2019-Q2 data before the tier existed, etc.)
      - '-' for not applicable
      - 'nfp' for not for publication (confidentialised)
      - '<10' / '<X' for cell-size suppression
    We treat all four as missing; downstream sums skip them.
    """
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in {"-", "nfp", "<10"} or s.startswith("<"):
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def load_tier_quarterly(path: Path) -> tuple[list[dict[str, Any]], str | None]:
    """Read the insured-persons block of sheet '4 HT by Product Tier'.

    The sheet contains two stacked blocks with identical layout:

      Rows ~1–29: "Total Gold/Silver/Bronze/Basic/Other hospital treatment **policies**"
      Row   ~30:  blank separator
      Rows ~31+:  "Total Gold/Silver/Bronze/Basic/Other hospital treatment **insured persons**"

    We locate the second block by searching column B for the "insured persons" group
    header, then read each tier's "Total Policies/Persons" summary column per quarter.

    Returns (records, caveat). Caveat explains the 2019-era legacy-product migration
    (the "Other" column shrinks from 5.57M to 0 between 2019-Q2 and 2020-Q2 as
    insurers migrated customers into the new Gold/Silver/Bronze/Basic taxonomy).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    if TIER_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{TIER_SHEET}' not found in {path.name}")
    ws = wb[TIER_SHEET]
    rows = list(ws.iter_rows(values_only=True))

    # Find the "insured persons" group header (column index 1).
    persons_header_idx: int | None = None
    for i, r in enumerate(rows):
        c1 = r[1] if len(r) > 1 else None
        if c1 and "insured persons" in str(c1).lower() and "gold" in str(c1).lower():
            persons_header_idx = i
            break
    if persons_header_idx is None:
        raise ValueError(
            f"Could not locate 'insured persons' block in sheet '{TIER_SHEET}' of {path.name}"
        )
    # Column headers sit one row below the group header; data starts two rows below.
    data_start = persons_header_idx + 2

    records: list[dict[str, Any]] = []
    for r in rows[data_start:]:
        if not r:
            continue
        q = r[0]
        if q is None:
            continue
        # Normalise quarter to APRA's convention (month-end ISO).
        if isinstance(q, datetime):
            qdate = q.date()
        elif isinstance(q, date):
            qdate = q
        else:
            continue
        # DoH labels quarters by the **first day of the reporting month** (e.g.
        # 2019-06-01 for Jun-2019 = APRA's 2019-06-30). Shift to quarter-end.
        y, m = qdate.year, qdate.month
        md = {3: "03-31", 6: "06-30", 9: "09-30", 12: "12-31"}.get(m)
        if not md:
            continue
        quarter_iso = f"{y}-{md}"
        if quarter_iso < SERIES_START_ISO:
            continue

        tiers: dict[str, float | None] = {}
        for tier, col in TIER_TOTAL_COLS.items():
            if col >= len(r):
                tiers[tier] = None
                continue
            tiers[tier] = _num(r[col])

        # Primary denominator: sum of published tier totals excluding the grand total.
        tier_vals = [tiers[t] for t in TIER_ORDER if tiers.get(t) is not None]
        tier_sum = sum(tier_vals) if tier_vals else None
        grand_total = tiers.get("grand_total")

        # Shares are computed against tier_sum (internally consistent, immune to the
        # small 2019-Q2 "Other vs grand total" rounding gap). grand_total is kept for
        # reconciliation against APRA.
        shares: dict[str, float | None] = {}
        for t in TIER_ORDER:
            v = tiers.get(t)
            shares[t] = (v / tier_sum) if (v is not None and tier_sum and tier_sum > 0) else None

        records.append(
            {
                "quarter": quarter_iso,
                "insured_persons": {t: tiers.get(t) for t in TIER_ORDER},
                "share": shares,
                "total_insured_persons": tier_sum,
                "grand_total_published": grand_total,
            }
        )

    records.sort(key=lambda x: x["quarter"])

    caveat = (
        "Insured persons with hospital treatment by product tier, from the Department "
        "of Health, Disability and Ageing 'Private Health Insurance Reform Data "
        "Quarterly Trends Report'. Gold/Silver/Bronze/Basic product tiers were "
        "introduced on 1 April 2019 and mandatory from 1 April 2020 — the 'Legacy' "
        "series captures pre-reform products that insurers migrated into the new "
        "taxonomy through 2019–Q1 2020 (after which it is zero). Read tier shares "
        "from ~2020-Q2 onward as the stable market picture."
    )
    return records, caveat


def reconcile(
    national: list[dict[str, Any]],
    expected_hospital: float | None,
    tolerance: float = 0.01,
) -> ReconciliationResult:
    if not national or expected_hospital is None:
        return ReconciliationResult(
            True,
            "hospital_insured_national_last_quarter",
            expected_hospital,
            None,
            None,
            tolerance,
            "No reconciliation target set — skipped.",
        )
    last = national[-1]["hospital_treatment"]["insured_persons"]
    if last is None:
        return ReconciliationResult(
            False,
            "hospital_insured_national_last_quarter",
            expected_hospital,
            None,
            None,
            tolerance,
            "Missing last-quarter hospital insured persons.",
        )
    diff = abs(last - expected_hospital)
    rel = diff / max(expected_hospital, 1.0)
    ok = rel <= tolerance
    return ReconciliationResult(
        ok,
        "hospital_insured_national_last_quarter",
        expected_hospital,
        last,
        diff,
        tolerance,
        "PASS" if ok else f"Relative diff {rel:.4f} exceeds tolerance {tolerance}.",
    )


def build_payload(
    xlsx_path: Path,
    mb_path: Path | None,
    erp_path: Path | None,
    tier_path: Path | None,
    expected_hospital: float | None,
    fail_on_reconcile: bool,
) -> dict[str, Any]:
    ht, gt = load_ht_gt(xlsx_path)
    national = build_national_quarterly_simple(ht, gt)
    by_state = build_state_quarterly_clean(ht)

    age_coverage: list[dict[str, Any]] = []
    age_notice: str | None = None
    if mb_path and mb_path.exists() and erp_path and erp_path.exists():
        age_df = load_age_cohort_ht(mb_path)
        erp_df = load_abs_erp(erp_path)
        age_coverage, age_notice = build_age_coverage_quarterly(age_df, erp_df)

    tier_quarterly: list[dict[str, Any]] = []
    tier_caveat: str | None = None
    if tier_path and tier_path.exists():
        tier_quarterly, tier_caveat = load_tier_quarterly(tier_path)

    rec = reconcile(national, expected_hospital)
    if not rec.passed and rec.message and "skipped" not in rec.message and fail_on_reconcile:
        raise SystemExit(f"Reconciliation failed: {rec.message} actual={rec.actual} expected={rec.expected}")

    # Secondary sanity check: sum of age-band insured persons should be close to the national
    # hospital insured-persons number from the Trends file for each common quarter.
    if age_coverage and national:
        nat_map = {
            n["quarter"]: (n["hospital_treatment"].get("insured_persons") or 0) for n in national
        }
        worst_rel = 0.0
        worst_q = None
        for pt in age_coverage:
            qnat = nat_map.get(pt["quarter"])
            if not qnat:
                continue
            total = sum(b["insured_persons"] for b in pt["bands"].values())
            rel = abs(total - qnat) / qnat
            if rel > worst_rel:
                worst_rel = rel
                worst_q = pt["quarter"]
        if worst_rel > 0.02:
            msg = (
                f"Age-band sum disagrees with national total by {worst_rel:.2%} "
                f"(worst quarter {worst_q}); check AgeCohort_HT coverage vs Membership Trends."
            )
            if fail_on_reconcile:
                raise SystemExit(msg)
            print(f"WARN: {msg}", file=sys.stderr)

    # Tertiary sanity check: DoH tier insured-persons total should reconcile to APRA
    # national hospital insured-persons within ~1% for each common quarter.
    tier_reconciliation: dict[str, Any] | None = None
    if tier_quarterly and national:
        nat_map = {
            n["quarter"]: (n["hospital_treatment"].get("insured_persons") or 0) for n in national
        }
        worst_rel = 0.0
        worst_q: str | None = None
        worst_apra: float | None = None
        worst_doh: float | None = None
        checked = 0
        for pt in tier_quarterly:
            qnat = nat_map.get(pt["quarter"])
            tot = pt.get("total_insured_persons")
            if not qnat or not tot:
                continue
            checked += 1
            rel = abs(tot - qnat) / qnat
            if rel > worst_rel:
                worst_rel = rel
                worst_q = pt["quarter"]
                worst_apra = float(qnat)
                worst_doh = float(tot)
        tol = 0.01
        ok = worst_rel <= tol
        tier_reconciliation = {
            "metric": "doh_tier_total_vs_apra_hospital_insured",
            "quarters_compared": checked,
            "worst_relative_diff": worst_rel,
            "worst_quarter": worst_q,
            "apra_at_worst": worst_apra,
            "doh_tier_sum_at_worst": worst_doh,
            "tolerance": tol,
            "passed": ok,
        }
        if not ok:
            msg = (
                f"DoH tier sum disagrees with APRA national hospital insured persons by "
                f"{worst_rel:.2%} at {worst_q} (APRA {worst_apra:.0f} vs DoH {worst_doh:.0f})."
            )
            if fail_on_reconcile:
                raise SystemExit(msg)
            print(f"WARN: {msg}", file=sys.stderr)

    now = datetime.now(timezone.utc).isoformat()
    last_q = national[-1]["quarter"] if national else None
    sources = [
        "Australian Prudential Regulation Authority (APRA) — Quarterly private health insurance membership trends.",
    ]
    if mb_path and mb_path.exists():
        sources.append(
            "Australian Prudential Regulation Authority (APRA) — Quarterly private health insurance membership and benefits (AgeCohort_HT)."
        )
    if erp_path and erp_path.exists():
        sources.append(
            "Australian Bureau of Statistics (ABS) — Estimated Resident Population, quarterly, by 5-year age band (ERP_Q SDMX dataflow)."
        )
    if tier_quarterly:
        sources.append(
            "Australian Government Department of Health, Disability and Ageing — "
            "Private Health Insurance Reform Data Quarterly Trends Report "
            "(sheet '4 HT by Product Tier', insured-persons block)."
        )

    meta: dict[str, Any] = {
        "data_as_of": last_q,
        "series_start_iso": SERIES_START_ISO,
        "etl_build_time_utc": now,
        "source_file": str(xlsx_path.name),
        "apra_methodology_change_iso": APRA_METHODOLOGY_CHANGE,
        "tier_reform_effective_iso": TIER_REFORM_ISO,
        "pop_denominator_note": (
            "Population denominators and published coverage shares are as in the APRA "
            "Membership Trends workbook for the national / state series. The age-band "
            "coverage chart uses ABS ERP by 5-year age band (persons, national) via the "
            "ABS Data API (dataflow ERP_Q). Age band '80+' aggregates APRA's 80-84, 85-89, "
            "90-94 and 95+ bands to match ABS ERP_Q's highest published band."
        ),
        "hospital_vs_general_note": (
            "Hospital treatment and general (extras) are separate insured-persons series. "
            "They are not additive as a headcount of distinct people (many people hold both)."
        ),
        "sources": sources,
    }
    if age_notice:
        meta["age_source_note"] = age_notice
    if tier_caveat:
        meta["tier_note"] = tier_caveat

    payload: dict[str, Any] = {
        "meta": meta,
        "reconciliation": asdict(rec),
        "national_quarterly": national,
        "jurisdiction_quarterly": by_state,
    }
    if age_coverage:
        payload["age_coverage_quarterly"] = age_coverage
    if tier_quarterly:
        payload["tier_quarterly"] = tier_quarterly
    if tier_reconciliation:
        payload["tier_reconciliation"] = tier_reconciliation
    return payload


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument(
        "--xlsx",
        type=Path,
        default=Path("data/raw/membership_trends_dec_2025.xlsx"),
    )
    p.add_argument(
        "--membership-benefits",
        type=Path,
        default=Path("data/raw/membership_benefits_dec_2025.xlsx"),
        help="APRA Quarterly Membership and Benefits workbook (AgeCohort_HT).",
    )
    p.add_argument(
        "--abs-erp",
        type=Path,
        default=Path("data/curated/abs_erp_national_5y.csv"),
        help="Tidy CSV of ABS ERP national 5-year age bands (see etl/fetch_abs_erp.py).",
    )
    p.add_argument(
        "--tier-xlsx",
        type=Path,
        default=Path("data/raw/phi_reform_trends.xlsx"),
        help=(
            "Department of Health 'Private Health Insurance Reform Data Quarterly "
            "Trends Report' workbook (sheet '4 HT by Product Tier')."
        ),
    )
    p.add_argument(
        "--out",
        type=Path,
        default=Path("public/data/dashboard.json"),
    )
    p.add_argument(
        "--expected-hospital-insured",
        type=float,
        default=None,
        help="If set, last-quarter national hospital insured persons must match within tolerance (default 1%).",
    )
    p.add_argument("--no-fail-reconcile", action="store_true")
    args = p.parse_args()
    if not args.xlsx.exists():
        print(f"Missing {args.xlsx}. Download the APRA 'Membership Trends' xlsx into data/raw/.", file=sys.stderr)
        sys.exit(1)
    ex = None
    if args.expected_hospital_insured is not None:
        ex = args.expected_hospital_insured
    elif os.environ.get("EXPECTED_HOSPITAL_INSURED"):
        ex = float(os.environ["EXPECTED_HOSPITAL_INSURED"])
    else:
        exp_path = Path(__file__).resolve().parent / "expected_reconciliation.json"
        if exp_path.exists():
            with exp_path.open(encoding="utf-8") as f:
                exp = json.load(f)
            v = exp.get("hospital_insured_national")
            if v is not None:
                ex = float(v)
    pl = build_payload(
        args.xlsx,
        args.membership_benefits if args.membership_benefits.exists() else None,
        args.abs_erp if args.abs_erp.exists() else None,
        args.tier_xlsx if args.tier_xlsx.exists() else None,
        ex,
        fail_on_reconcile=not args.no_fail_reconcile,
    )
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(pl, indent=2), encoding="utf-8")
    n_age = len(pl.get("age_coverage_quarterly", []))
    n_tier = len(pl.get("tier_quarterly", []))
    print(
        f"Wrote {args.out} ({len(pl['national_quarterly'])} quarters national, "
        f"{n_age} age-coverage quarters, {n_tier} tier quarters)."
    )


if __name__ == "__main__":
    main()
